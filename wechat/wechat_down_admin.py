import requests
import time
import json
from openpyxl import Workbook
import random
url = "https://mp.weixin.qq.com/cgi-bin/appmsg"
Cookie = ""
headers = {
    "Cookie": Cookie,
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36 Edg/106.0.1370.52",
}
# https://github.com/YPstar-yes/Crawl-WeChat-articles/blob/master/weixin.py
token = ""#公众号 people/bai-ri-meng-yu-shi-76 
fakeid = "MjM5NjM4MDAxMg=="#公众号id
type = '9'
data1 = {
    "token": token,
    "lang": "zh_CN",
    "f": "json",
    "ajax": "1",
    "action": "list_ex",
    "begin": "0",
    "count": "4",
    "query": "",
    "fakeid": fakeid,
    "type": type,
}
def getMoreInfo(link):
    # 获得mid,_biz,idx,sn
    mid = link.split("&")[1].split("=")[1]
    idx = link.split("&")[2].split("=")[1]
    sn = link.split("&")[3].split("=")[1]
    _biz = link.split("&")[0].split("_biz=")[1]
    pass_ticket = ""#抓包
    appmsg_token = ""#
    url = "http://mp.weixin.qq.com/mp/getappmsgext"#获取详情页
    phoneCookie = ""
    headers = {
        "Cookie": phoneCookie,
        "User-Agent": ""
    }
    data = {
        "is_only_read": "1",
        "is_temp_url": "0",
        "appmsg_type": "9",
        'reward_uin_count': '0'
    }
    params = {
        "__biz": _biz,
        "mid": mid,
        "sn": sn,
        "idx": idx,
        "key": "",
        "pass_ticket": pass_ticket,
        "appmsg_token": appmsg_token,
        "uin": "",
        "wxtoken": "777",
    }
    requests.packages.urllib3.disable_warnings()
    content = requests.post(url, headers=headers, data=data, params=params).json()
    # print(content["appmsgstat"]["read_num"], content["appmsgstat"]["like_num"])
    try:
        readNum = content["appmsgstat"]["read_num"]
        print("阅读数:"+str(readNum))
    except:
        readNum = 0
    try:
        likeNum = content["appmsgstat"]["like_num"]
        print("点赞数:"+str(likeNum))
    except:
        likeNum = 0
    try:
        old_like_num = content["appmsgstat"]["old_like_num"]
        print("在看数:"+str(old_like_num))
    except:
        old_like_num = 0
    time.sleep(3) #
    return readNum, likeNum,old_like_num
def getAllInfo(url):
    messageAllInfo = []
    for i in range(33):
        begin = i * 4
        data1["begin"] = begin
        requests.packages.urllib3.disable_warnings()
        content_json = requests.get(url, headers=headers, params=data1, verify=False).json()
        time.sleep(random.randint(1, 10))
        if "app_msg_list" in content_json:
            for item in content_json["app_msg_list"]:
                spider_url = item['link']
                readNum, likeNum,old_like_num = getMoreInfo(spider_url)
                info = {
                    "title": item['title'],
                    "url": item['link'],
                    "readNum": readNum,
                    "likeNum": likeNum,
                    "old_like_num":old_like_num
                }
                messageAllInfo.append(info)
    return messageAllInfo
def main():
    f = Workbook()
    sheet = f.active
    sheet.cell(row=1, column=1).value = 'title'
    sheet.cell(row=1, column=2).value = 'url'
    sheet.cell(row=1, column=3).value = 'readNum'
    sheet.cell(row=1, column=4).value = 'likeNum'
    sheet.cell(row=1, column=5).value = 'old_like_num'
    messageAllInfo = getAllInfo(url)
    print(messageAllInfo)
    print(len(messageAllInfo))
    for i in range(1, len(messageAllInfo)+1):
        sheet.cell(row=i + 1, column=1).value = messageAllInfo[i - 1]['title']
        sheet.cell(row=i + 1, column=2).value = messageAllInfo[i - 1]['url']
        sheet.cell(row=i + 1, column=3).value = messageAllInfo[i - 1]['readNum']
        sheet.cell(row=i + 1, column=4).value = messageAllInfo[i - 1]['likeNum']
        sheet.cell(row=i + 1, column=5).value = messageAllInfo[i - 1]['old_like_num']
    f.save(u'公众号文章列表.xls')  #
if __name__ == '__main__':
    main()